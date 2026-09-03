"""任务执行器（核心调度）

驱动 Playwright 浏览器执行规则步骤，
集成弹窗拦截、新窗口接管、随机延迟、智能自愈、截图与报告。
"""

from __future__ import annotations

import asyncio
import csv
import json
import random
import re
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from config import SCREENSHOT_DIR, REPORT_DIR, EXPORT_DIR, SPEED_MODES, ensure_dirs
from llm_client import LLMClient, LLMError
from pii import PIIGateway
from review import add_review
from rule_engine import RuleEngine, CONTROL_FLOW_TYPES
from self_healing import SelfHealing, HealingContext, ElementNotFoundError
from template_engine import TemplateEngine

try:
    from playwright.async_api import (
        Browser,
        BrowserContext,
        Page,
        async_playwright,
    )
except ImportError:  # pragma: no cover
    Browser = BrowserContext = Page = Any  # type: ignore
    async_playwright = None  # type: ignore


def _is_empty(v: Any) -> bool:
    """判断单元格值是否为空（None / 空串 / 纯空白）。"""
    return v is None or (isinstance(v, str) and v.strip() == "")


def _normalize_date(v: Any) -> str:
    """把各种奇葩日期格式统一成 YYYY-MM-DD；无法解析返回空串。"""
    from datetime import date, datetime

    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if not isinstance(v, str):
        return ""
    s = v.strip()
    if not s:
        return ""
    # 常见分隔符（含中文年月日）统一替换为 - 再解析
    cand = re.sub(r"[年月日./.]", "-", s)
    cand = re.sub(r"[-]+", "-", cand).strip("-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return datetime.strptime(cand[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


def _detect_numeric_outliers(batch: List[Dict[str, Any]], skip_cols: Iterable[str]) -> Dict[int, List[str]]:
    """确定性数值离群检测：某数值列的值远超同批中位数时标记异常（进人工审核兜底）。

    返回 {行下标: [异常描述]}。姓名/标识列与索引列跳过；少于 3 条或中位数为 0 时不判。
    """
    from statistics import median

    skip = set(skip_cols or ())
    numeric_cols: List[str] = []
    for r in batch:
        for k, v in r.items():
            if k in skip or k in ("index", "confidence"):
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool) and k not in numeric_cols:
                numeric_cols.append(k)
    flags: Dict[int, List[str]] = {}
    for k in numeric_cols:
        vals = [r.get(k) for r in batch if isinstance(r.get(k), (int, float))]
        if len(vals) < 3:
            continue
        med = median(vals)
        if med == 0:
            continue
        for i, r in enumerate(batch):
            v = r.get(k)
            if isinstance(v, (int, float)) and abs(v) > abs(med) * 10:
                flags.setdefault(i, []).append(f"数值异常：{k}={v}，远超同批中位数{med}")
    return flags


class BrowserManager:
    """浏览器生命周期管理。"""

    def __init__(self):
        self._pw = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        # attach 接管模式下，连接的是用户自己的调试浏览器（登录态/数据都在里面）。
        # 关闭时绝不能调用 browser.close() 把它关掉，否则：①用户浏览器被关；
        # ②浏览器进程被终结后调试端口虽在监听，但 context 已失效，下次执行 new_page 会报
        #    "Target page, context or browser has been closed"，导致任务 steps=0 直接失败。
        self._attached = False

    async def start(self, headless: bool = True) -> Browser:
        self._pw = await async_playwright().start()
        self.browser = await self._pw.chromium.launch(headless=headless)
        self.context = await self.browser.new_context()
        return self.browser

    async def new_page(self) -> Page:
        return await self.context.new_page()

    async def start_attach(self, cdp_url: str) -> Browser:
        """连接到一个已以调试模式启动的浏览器（Edge/Chrome/360）。"""
        self._pw = await async_playwright().start()
        self.browser = await self._pw.chromium.connect_over_cdp(cdp_url)
        contexts = self.browser.contexts
        # 优先沿用已有上下文（用户已登录的浏览器），无则新建
        self.context = contexts[0] if contexts else await self.browser.new_context()
        self._attached = True
        return self.browser

    async def close(self) -> None:
        try:
            if self._attached:
                # attach 接管模式：浏览器归用户所有（登录态/数据在其中），且 executor 为
                # 模块级单例、会被连续复用。绝不能 browser.close()（会关掉用户浏览器），
                # 也不能 _pw.stop()（会断开 CDP 连接并使后续新连接落入 closed 态，
                # 表现为 new_page: context has been closed）。直接保留连接即可。
                return
            if self.browser:
                await self.browser.close()
            if self._pw:
                await self._pw.stop()
        except Exception:
            pass
        finally:
            if not self._attached:
                # 内置模式：清空引用，否则下次执行时 run() 会误判 context 仍可用，
                # 直接对已关闭的 context 调 new_page()，报 "Target page, context or
                # browser has been closed"，任务 steps=0 失败。
                self._pw = None
                self.browser = None
                self.context = None


class PageOperator:
    """页面操作封装，统一处理自愈定位与动作执行。"""

    def __init__(self, healing: SelfHealing):
        self.healing = healing
        # 任务内标签栈（栈底=起始列表页）：点击在新标签打开页面时记录来源页，
        # close_tab 关闭当前标签后回退到「打开它之前的页面」。
        self._tab_stack: List[Page] = []

    def _push_tab_stack(self, page: Page) -> None:
        """记录「新标签页是从哪个页面打开的」，供 close_tab 回退。"""
        stack = getattr(self, "_tab_stack", None)
        if stack is None:
            return
        if stack and stack[-1] is page:
            return  # 避免重复压入同一个页面
        stack.append(page)

    def _pop_tab_stack(self, old_page: Optional[Page]) -> Optional[Page]:
        """关闭标签后从栈里找「打开它之前的最近有效页面」；找不到返回 None。"""
        stack = getattr(self, "_tab_stack", None)
        if not stack:
            return None
        # 先移除旧当前页（可能在栈中），再找最近的有效页
        while stack:
            top = stack[-1]
            if top is old_page:
                stack.pop()
                continue
            try:
                if top.is_closed() is False:
                    return top
            except Exception:
                pass
            stack.pop()
        return None

    async def _resolve_locator(
        self, page: Page, selector: str, intent: str, ctx: HealingContext
    ):
        """通过自愈机制定位元素，返回 locator 或坐标。"""
        result = await self.healing.locate(selector, intent, page, ctx)
        if result.method == "coordinate":
            return ("coordinate", result.value)
        return ("locator", page.locator(result.value).first)

    async def _capture_clicked(self, page: Page, kind: str, val: Any) -> Optional[dict]:
        """读取将要点中的元素：可见文字 + 真实链接。失败返回 None。"""
        try:
            if kind == "coordinate":
                x, y = val
                return await page.evaluate(
                    """(xy) => {
                        const el = document.elementFromPoint(
                            Math.round(innerWidth * xy.x), Math.round(innerHeight * xy.y));
                        if (!el) return null;
                        const link = el.closest('a[href]');
                        return { text: (el.innerText || el.textContent || '').trim().slice(0, 40),
                                 href: link ? link.href : null };
                    }""",
                    {"x": x, "y": y},
                )
            return await val.evaluate(
                "el => { const link = el.closest('a[href]'); "
                "return { text: (el.innerText || el.textContent || '').trim().slice(0, 40), "
                "href: link ? link.href : null }; }"
            )
        except Exception:
            return None

    async def click(self, page: Page, selector: str, intent: str = "", ctx: Optional[HealingContext] = None,
                    current_page: Optional[List[Page]] = None) -> bool:
        ctx = ctx or HealingContext()
        kind, val = await self._resolve_locator(page, selector, intent or "点击元素", ctx)
        # 点击前记录真实点击目标（被点元素的可见文字与真实链接），供历史/报告展示，
        # 让用户清楚"这次真实点到了什么"，而不是只看到跳转后的网址
        ctx.clicked = await self._capture_clicked(page, kind, val)
        if kind == "coordinate":
            await SelfHealing.click_by_coordinate(page, *val)
            return True
        # 低信任点击（意图无有效文字标签，如只有 CSS 选择器）：先做视觉复核，
        # 防止把「热门」点到相邻的「娱乐」；明确判定不是目标时改用全页视觉定位坐标。
        if kind == "locator" and intent and self.healing.is_low_trust(intent):
            confirm = await self.healing.visual_confirm(page, intent, val)
            if confirm is False:
                ctx.log("视觉复核判定命中元素不是目标，改用全页视觉定位")
                coord = await self.healing.visual_locate(intent, page)
                if coord:
                    await SelfHealing.click_by_coordinate(page, *coord)
                    return True
                raise ElementNotFoundError(
                    f"视觉复核未通过且视觉定位失败，已取消点击以防点错：{intent}"
                )
        # 目标是链接时，记住其绝对 href 与 target
        href = None
        new_tab = False
        try:
            href, new_tab = await val.evaluate(
                "el => { const c = el.closest('a[href]'); "
                "return c ? [c.href, c.target === '_blank'] : [null, false]; }"
            )
        except Exception:
            pass
        before = page.url
        # 捕获点击可能打开的新标签页（target="_blank"），点击成功后若确实开了新标签页就跟随它。
        popup = None
        try:
            async with page.expect_popup(timeout=9000) as popup_info:
                # 正常点击，缩短默认超时；元素不稳定（轮播图动画）时快速降级
                try:
                    await val.click(timeout=8000)
                except Exception:
                    ctx.log("元素不稳定/被遮挡，降级为强制点击")
                    try:
                        await val.click(force=True, timeout=5000)
                    except Exception:
                        ctx.log("强制点击失败，降级为 JS 直接点击")
                        try:
                            await val.evaluate("el => el.click()")
                        except Exception:
                            pass
            popup = await popup_info.value
        except Exception:
            # 未打开新标签页（普通点击）时，expect_popup 超时，走下方跳转验证
            popup = None

        # 1) 点击打开了新标签页：跟随它。B 站等会给链接追加 spm 追踪参数
        #    （?spm_id_from=...）导致 popup.url != href，故只要新标签页是真实页面就跟随。
        if popup is not None and current_page is not None:
            try:
                await popup.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
            try:
                purl = popup.url
            except Exception:
                purl = ""
            if purl and purl != "about:blank":
                # 记录来源页，供后续 close_tab 回退到「打开该标签前的页面」
                self._push_tab_stack(current_page[0])
                ctx.log(f"点击在新标签页打开了 {purl}，已切换为该标签页")
                current_page[0] = popup
                return True
            # 新标签页仍在加载（URL 未就绪），也先跟随，后续步骤会等待加载
            self._push_tab_stack(current_page[0])
            ctx.log("点击打开了新标签页（仍在加载），已切换为该标签页")
            current_page[0] = popup
            return True

        # 统一判断 URL 是否跳转的辅助函数（去掉查询参数与锚点，避免 spm 追踪参数干扰）
        def _norm(u: str) -> str:
            return (u or "").split("#")[0].split("?")[0].rstrip("/") or u or ""

        # 2) 当前页跳转验证：链接被弹窗/遮罩拦截时「点击成功但无反应」。
        #    target="_blank" 且未捕获到 popup 时跳过强制导航，避免 goto 造成「打开两个相同网页」。
        if href and href != before and not new_tab:
            navigated = False
            try:
                await page.wait_for_url(lambda u: u != before, timeout=8000)
                navigated = True
            except Exception:
                navigated = False
            if not navigated:
                # 目标可能已被其它标签页打开（响应慢导致跳转验证超时），切换到它而不是重复 goto
                target = None
                try:
                    if _norm(page.url) != _norm(before):
                        target = page
                    else:
                        for p in page.context.pages:
                            if p is not page:
                                try:
                                    if _norm(p.url) == _norm(href):
                                        target = p
                                        break
                                except Exception:
                                    pass
                except Exception:
                    pass
                if target is not None:
                    ctx.log(f"目标页面已由其他标签页打开（{target.url}），切换过去")
                    if current_page is not None:
                        current_page[0] = target
                else:
                    ctx.log(f"点击链接后页面未跳转（可能被遮罩拦截），直接导航到 {href}")
                    try:
                        await page.goto(href, wait_until="domcontentloaded", timeout=15000)
                    except Exception:
                        pass
        # 3) target="_blank" 但 expect_popup 未捕获（如 JS 降级点击）时，
        #    若发现其它标签页已打开目标地址，也切换过去。
        elif href and new_tab and current_page is not None:
            try:
                for p in page.context.pages:
                    if p is not page:
                        try:
                            if _norm(p.url) == _norm(href):
                                ctx.log(f"目标页面已由其他标签页打开（{p.url}），切换过去")
                                current_page[0] = p
                                break
                        except Exception:
                            pass
            except Exception:
                pass
        return True

    async def input_text(self, page: Page, selector: str, value: str, speed: str = "normal",
                         intent: str = "", ctx: Optional[HealingContext] = None) -> bool:
        ctx = ctx or HealingContext()
        kind, val = await self._resolve_locator(page, selector, intent or "输入内容", ctx)
        if kind == "coordinate":
            await SelfHealing.click_by_coordinate(page, *val)
            target = page
        else:
            target = val
            await val.click()
        # 模拟打字
        lo, hi = SPEED_MODES.get(speed, SPEED_MODES["normal"])
        for ch in str(value):
            await target.keyboard.type(ch, delay=random.randint(30, 90))
        return True

    async def select(self, page: Page, selector: str, value: str, intent: str = "",
                     ctx: Optional[HealingContext] = None) -> bool:
        ctx = ctx or HealingContext()
        kind, val = await self._resolve_locator(page, selector, intent or "选择下拉项", ctx)
        if kind == "coordinate":
            await SelfHealing.click_by_coordinate(page, *val)
            return True
        await val.select_option(value)
        return True

    async def extract(self, page: Page, selector: str, extract_type: str = "text",
                      attribute: str = "", intent: str = "", ctx: Optional[HealingContext] = None) -> Any:
        ctx = ctx or HealingContext()
        # 提取动作不要求元素可见（如 <title>、隐藏字段），先尝试直接定位
        locator = page.locator(selector).first
        try:
            if await locator.count() == 0:
                raise Exception("no element")
        except Exception:
            # 直接定位失败，走自愈降级
            kind, val = await self._resolve_locator(page, selector, intent or "提取数据", ctx)
            if kind == "coordinate":
                return None
            locator = val
        if extract_type == "text":
            # text_content 可提取隐藏元素文本；title 等标签 inner_text 为空
            text = await locator.text_content()
            return (text or "").strip()
        if extract_type == "attribute":
            return await locator.get_attribute(attribute or "value")
        if extract_type == "inner_html":
            return await locator.inner_html()
        if extract_type == "value":
            return await locator.input_value()
        if extract_type == "count":
            return await locator.count()
        return (await locator.text_content() or "").strip()


class TaskExecutor:
    """任务执行器。"""

    def __init__(self):
        self.manager = BrowserManager()
        self.healing = SelfHealing()
        self.operator = PageOperator(self.healing)

    async def run(
        self,
        task_config: Dict[str, Any],
        on_step_log=None,
        on_screenshot=None,
        exec_id: str = "",
        headless: bool = True,
        stop_event: Optional[asyncio.Event] = None,
        attach_url: Optional[str] = None,
    ) -> Dict[str, Any]:
        """执行完整任务，返回执行结果摘要。"""
        ensure_dirs()
        stop_event = stop_event or asyncio.Event()
        speed = task_config.get("speed_mode", "normal")
        steps = RuleEngine.normalize_steps(task_config.get("steps", []))
        context_vars: Dict[str, Any] = {}
        context_vars["__task_name"] = task_config.get("task_name", "task")[:40]
        step_logs: List[Dict[str, Any]] = []
        start_time = time.time()

        screenshot_dir = SCREENSHOT_DIR / (task_config.get("task_name", "task")[:20] or "task") / (exec_id or str(int(start_time)))
        screenshot_dir.mkdir(parents=True, exist_ok=True)

        try:
            if self.manager.context is None:
                if attach_url:
                    await asyncio.wait_for(self.manager.start_attach(attach_url), timeout=15)
                else:
                    await asyncio.wait_for(self.manager.start(headless=headless), timeout=30)
        except asyncio.TimeoutError:
            await self.manager.close()
            raise RuntimeError(
                "浏览器启动/连接超时：内置浏览器 30 秒未就绪，或接管模式未连接到调试浏览器（默认端口 9222）"
            )
        page = await self.manager.new_page()
        await self._register_page_handlers(page)

        current_page: List[Page] = [page]
        # 记录任务自己的「起始列表页」引用：任务第一步 open 的目标页。
        # attach 接管模式下浏览器可能开着用户其它标签页（如应用前端页），
        # close_tab 等需要「回到列表页」时不能从所有标签页里随便挑，
        # 否则可能切到用户的页面导致后续步骤全部定位失败。
        self._anchor_page = page
        # 任务内标签栈（栈底=起始列表页）：每次点击在新标签打开页面时记录来源页，
        # close_tab 关闭当前标签后回到「打开它之前的页面」（如热门页→视频页→
        # 关闭视频页回热门页），而不是硬回锚点页，避免后续步骤在错误的页面执行。
        self.operator._tab_stack = [page]

        try:
            # 建立 step_id -> 数组下标 映射，供 goto/if 跳转
            id_to_index: Dict[int, int] = {}
            for idx, s in enumerate(steps):
                id_to_index[int(s.get("step_id", idx + 1))] = idx

            # 程序指针 + 最大步数保护（防止死循环）
            pc = 0
            max_steps = max(1000, len(steps) * 200)
            executed = 0

            while pc < len(steps):
                if stop_event.is_set():
                    break
                if executed >= max_steps:
                    step_logs.append({
                        "step_id": steps[pc].get("step_id"),
                        "action_type": "loop_guard",
                        "status": "failed",
                        "error_info": f"超过最大执行步数 {max_steps}，疑似死循环已终止",
                    })
                    break

                step = steps[pc]
                action = RuleEngine.resolve_params(step.get("action", {}), context_vars)
                condition = step.get("condition", {}) or {}
                step_id = step.get("step_id", len(step_logs) + 1)
                atype = action.get("type", "")
                intent = self._build_intent(action, step.get("note") or "")
                ctx = HealingContext()

                # ---- 控制流动作：不操作页面，只改 pc ----
                if atype in CONTROL_FLOW_TYPES:
                    jump = await self._eval_control_flow(
                        current_page[0], action, context_vars
                    )
                    try:
                        page_url = current_page[0].url
                    except Exception:
                        page_url = None
                    log = {
                        "step_id": step_id,
                        "action_type": atype,
                        "target_element": self._control_target(action),
                        "status": "success",
                        "duration_ms": 0,
                        "healing_actions": [],
                        "extracted": jump["remark"],
                        "page_url": page_url,
                    }
                    step_logs.append(log)
                    if on_step_log:
                        on_step_log(log)
                    if jump["to"] is None:
                        break  # 结束执行
                    target_idx = id_to_index.get(jump["to"])
                    if target_idx is None:
                        log["status"] = "failed"
                        log["error_info"] = f"跳转目标 step_id {jump['to']} 不存在"
                        break
                    pc = target_idx
                    executed += 1
                    continue

                # ---- 普通动作 ----
                t0 = time.time()
                before_path = await self._shot(current_page[0], screenshot_dir, f"{step_id}_before")
                try:
                    # 条件等待
                    if not await self._wait_condition(current_page[0], condition):
                        raise ElementNotFoundError(f"条件未满足: {condition.get('type')}")

                    # 执行动作
                    result = await self._execute_action(
                        current_page, action, speed, intent, ctx, stop_event, context_vars
                    )
                    # 随机延迟（模拟人类）
                    lo, hi = SPEED_MODES.get(speed, SPEED_MODES["normal"])
                    await asyncio.sleep(random.uniform(lo, hi))

                    # 提取变量
                    save_as = action.get("save_as")
                    if save_as and result is not None:
                        context_vars[save_as] = result

                    duration = int((time.time() - t0) * 1000)
                    after_path = await self._shot(current_page[0], screenshot_dir, f"{step_id}_after")
                    # 记录「该步骤执行后」所在页面
                    try:
                        page_url = current_page[0].url
                    except Exception:
                        page_url = None
                    log = {
                        "step_id": step_id,
                        "action_type": atype,
                        "target_element": action.get("selector") or action.get("url") or "",
                        "status": "success",
                        "duration_ms": duration,
                        "screenshot_before": str(before_path) if before_path else None,
                        "screenshot_after": str(after_path) if after_path else None,
                        "healing_actions": ctx.actions,
                        "clicked_info": ctx.clicked,
                        "extracted": result,
                        "page_url": page_url,
                    }
                    step_logs.append(log)
                    if on_step_log:
                        on_step_log(log)
                except Exception as e:
                    duration = int((time.time() - t0) * 1000)
                    err_text = str(e) or repr(e)
                    try:
                        page_url = current_page[0].url
                    except Exception:
                        page_url = None
                    log = {
                        "step_id": step_id,
                        "action_type": atype,
                        "target_element": action.get("selector") or action.get("url") or "",
                        "status": "failed",
                        "duration_ms": duration,
                        "healing_actions": ctx.actions,
                        "error_info": err_text,
                        "page_url": page_url,
                    }
                    step_logs.append(log)
                    if on_step_log:
                        on_step_log(log)
                    # 失败不中断，继续后续步骤，但记录错误
                pc += 1
                executed += 1
        finally:
            await self.manager.close()

        total_ms = int((time.time() - start_time) * 1000)
        failed = [l for l in step_logs if l["status"] == "failed"]
        status = "success" if not failed else ("failed" if len(failed) == len(step_logs) else "partial")
        return {
            "status": status,
            "duration_ms": total_ms,
            "steps": step_logs,
            "error_msg": "; ".join(l.get("error_info", "") for l in failed) if failed else None,
        }

    # ---- 内部实现 ----
    def _build_intent(self, action: Dict[str, Any], note: str = "") -> str:
        atype = action.get("type")
        desc = action.get("text") or action.get("selector") or action.get("url") or ""
        map_intent = {
            "click": f"点击元素「{desc}」",
            "input": f"在输入框「{desc}」输入内容",
            "select": f"选择「{desc}」中的选项",
            "extract": f"提取「{desc}」的数据",
        }
        intent = map_intent.get(atype, f"执行{atype}操作：{desc}")
        # 附上大白话说明，让第3/4层（AI 分析 DOM / 视觉定位）拿到语义上下文，
        # 避免只有 CSS 选择器时模型只能瞎猜（如点「热门」却落到相邻的「娱乐」）。
        if note:
            intent = f"{intent}；说明：{note}"
        return intent

    async def _eval_control_flow(
        self, page: Page, action: Dict[str, Any], context_vars: Dict[str, Any]
    ) -> Dict[str, Any]:
        """求值控制流动作，返回 {"to": <step_id|None>, "remark": str}。

        - goto: 无条件跳转到 target
        - if_text: 页面文本包含指定关键词则跳 goto_if_found，否则 goto_if_not
        - if_element: 元素存在则跳 goto_if_found，否则 goto_if_not
        to 为 None 表示结束执行。
        """
        atype = action.get("type")

        if atype == "goto":
            return {"to": action.get("target"), "remark": f"跳转到步骤 {action.get('target')}"}

        if atype == "if_text":
            text = str(action.get("text", ""))
            try:
                body_text = await page.evaluate("() => document.body.innerText || ''")
            except Exception:
                body_text = ""
            found = text and (text in body_text)
            target = action.get("goto_if_found") if found else action.get("goto_if_not")
            label = "命中" if found else "未命中"
            return {"to": target, "remark": f"if_text 关键词「{text}」{label} → 跳转 {target}"}

        if atype == "if_element":
            selector = action.get("selector", "")
            try:
                found = await page.locator(selector).first.count() > 0
            except Exception:
                found = False
            target = action.get("goto_if_found") if found else action.get("goto_if_not")
            label = "存在" if found else "不存在"
            return {"to": target, "remark": f"if_element 「{selector}」{label} → 跳转 {target}"}

        if atype == "if_var":
            var_name = action.get("var", "")
            expected = str(action.get("value", ""))
            op = action.get("op", "equals")  # equals/contains/not_equals/not_contains/less/less_equals/greater/greater_equals
            actual = context_vars.get(var_name)
            actual_s = "" if actual is None else str(actual)
            # 数字比较（用于循环计数，如 i > 3 则结束）
            if op in ("less", "less_equals", "greater", "greater_equals"):
                try:
                    av, ev = float(actual_s), float(expected)
                except (TypeError, ValueError):
                    av = ev = float("nan")
                if op == "less":
                    found = av < ev
                elif op == "less_equals":
                    found = av <= ev
                elif op == "greater":
                    found = av > ev
                else:
                    found = av >= ev
            elif op == "equals":
                found = actual_s == expected
            elif op == "not_equals":
                found = actual_s != expected
            elif op == "not_contains":
                found = expected not in actual_s
            else:  # contains
                found = expected in actual_s
            target = action.get("goto_if_found") if found else action.get("goto_if_not")
            label = "命中" if found else "未命中"
            return {
                "to": target,
                "remark": f"if_var {var_name}={actual_s!r} {op} {expected!r} {label} → 跳转 {target}",
            }

        return {"to": None, "remark": f"未知控制流动作 {atype}"}

    def _control_target(self, action: Dict[str, Any]) -> str:
        atype = action.get("type", "")
        if atype == "goto":
            return f"target={action.get('target')}"
        if atype == "if_text":
            return f"text={action.get('text')} → 有[{action.get('goto_if_found')}]/无[{action.get('goto_if_not')}]"
        if atype == "if_element":
            return f"selector={action.get('selector')} → 有[{action.get('goto_if_found')}]/无[{action.get('goto_if_not')}]"
        if atype == "if_var":
            return f"var {action.get('var')} {action.get('op', 'contains')} {action.get('value')} → 是[{action.get('goto_if_found')}]/否[{action.get('goto_if_not')}]"
        if atype == "set_var":
            return f"var {action.get('var')} {action.get('op', 'set')} {action.get('value')}"
        return atype

    async def _register_page_handlers(self, page: Page) -> None:
        # 弹窗自动关闭
        page.on("dialog", lambda d: asyncio.ensure_future(self._dismiss_dialog(d)))

    async def _dismiss_dialog(self, dialog) -> None:
        try:
            await dialog.dismiss()
        except Exception:
            pass

    async def _wait_condition(self, page: Page, condition: Dict[str, Any]) -> bool:
        ctype = condition.get("type", "always")
        timeout = condition.get("timeout", 10000)
        if ctype == "always":
            return True
        if ctype == "page_load":
            try:
                await page.wait_for_load_state("domcontentloaded", timeout=timeout)
                return True
            except Exception:
                return True
        if ctype == "element_visible":
            sel = condition.get("selector")
            try:
                await page.locator(sel).first.wait_for(state="visible", timeout=timeout)
                return True
            except Exception:
                return False
        if ctype == "text_appears":
            text = condition.get("text", "")
            try:
                await page.get_by_text(text).first.wait_for(state="visible", timeout=timeout)
                return True
            except Exception:
                return False
        return True

    async def _execute_action(
        self, current_page: List[Page], action: Dict[str, Any], speed: str,
        intent: str, ctx: HealingContext, stop_event: asyncio.Event,
        context_vars: Dict[str, Any],
    ) -> Any:
        page = current_page[0]
        atype = action.get("type")
        if atype == "open":
            # 巨潮等网站偶发 ERR_EMPTY_RESPONSE/断连，goto 失败自动重试，避免停在空白页
            await self._safe_goto(page, action.get("url", ""))
            await self._settle(page)
            return None
        if atype == "click":
            await self.operator.click(page, action.get("selector", ""), intent, ctx, current_page)
            page = current_page[0]  # 点击可能打开了新标签页，跟随它
            await self._settle(page)
            return None
        if atype == "input":
            await self.operator.input_text(page, action.get("selector", ""),
                                           action.get("value", ""), speed, intent, ctx)
            return None
        if atype == "select":
            await self.operator.select(page, action.get("selector", ""),
                                       action.get("value", ""), intent, ctx)
            return None
        if atype == "extract":
            return await self.operator.extract(
                page, action.get("selector", ""),
                action.get("extract_type", "text"),
                action.get("attribute", ""), intent, ctx
            )
        if atype == "ocr":
            # OCR 读图：整页截图或某元素截图交给多模态模型识别文字
            return await self._action_ocr(page, action)
        if atype == "llm_extract":
            # 结构化抽取：取到页面/变量文字后，让 LLM 按 fields 抽成 JSON
            return await self._action_llm_extract(page, action, context_vars, ctx)
        if atype == "export":
            return self._action_export(action, context_vars, task_name=context_vars.get("__task_name", ""))
        if atype == "read_excel":
            result = self._action_read_excel(action)
            # 未显式 save_as 时自动写入 __table，供 data_clean/llm_summarize/export 流水线使用
            if not action.get("save_as"):
                context_vars["__table"] = result
            return result
        if atype == "read_csv":
            result = self._action_read_csv(action)
            if not action.get("save_as"):
                context_vars["__table"] = result
            return result
        if atype == "ocr_to_json":
            return await self._action_ocr_to_json(page, action, context_vars, ctx)
        if atype == "data_clean":
            return self._action_data_clean(action, context_vars)
        if atype == "llm_summarize":
            return await self._action_llm_summarize(action, context_vars, ctx)
        if atype == "wait":
            try:
                await asyncio.sleep(float(action.get("value", 1)))
            except (TypeError, ValueError):
                await asyncio.sleep(1)
            return None
        if atype == "scroll":
            await page.mouse.wheel(0, action.get("amount", 300) or 300)
            return None
        if atype == "hover":
            kind, val = await self.operator._resolve_locator(
                page, action.get("selector", ""), intent, ctx
            )
            if kind == "locator":
                await val.hover()
            return None
        if atype == "press_key":
            await page.keyboard.press(action.get("keys", "Enter"))
            return None
        if atype == "upload":
            kind, val = await self.operator._resolve_locator(
                page, action.get("selector", ""), intent or "上传文件", ctx
            )
            if kind == "locator":
                await val.set_input_files(action.get("files", []))
            return None
        if atype == "reload":
            try:
                await page.reload(wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            await self._settle(page)
            return None
        if atype == "back":
            try:
                await page.go_back(wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            await self._settle(page)
            return None
        if atype == "forward":
            try:
                await page.go_forward(wait_until="domcontentloaded", timeout=15000)
            except Exception:
                pass
            await self._settle(page)
            return None
        if atype == "close_tab":
            # 关闭标签页。两种模式：
            #  1) 指定目标（close_target）：按页面标题/网址关键词匹配要关闭的标签页，
            #     用于「打开番剧+综艺 → 分别关掉它们」。匹配不到时不关闭、不报错。
            #  2) 默认（无 close_target）：关闭「当前标签页」（如视频页），回到起始列表页。
            #     只应关闭临时打开的标签页，绝不可关掉任务起始列表页（否则失去回退页面，
            #     并可能反复 new_page 空白页）。
            old = page
            close_target = (action.get("close_target") or "").strip()

            if close_target:
                # 模式1：关闭匹配的标签页
                target_close = None
                all_pages = list(old.context.pages)
                for p in all_pages:
                    try:
                        u = (p.url or "")
                    except Exception:
                        u = ""
                    try:
                        if p.is_closed() is False:
                            title = await p.title() or ""
                        else:
                            title = ""
                    except Exception:
                        title = ""
                    if close_target in u or close_target in title:
                        target_close = p
                        break
                if target_close is not None:
                    try:
                        await target_close.close()
                    except Exception:
                        pass
                # 关闭后选择新的当前页：优先任务起始列表页
                await self._close_pick_current(all_pages, target_close, current_page)
                await self._settle(current_page[0])
                return None

            # 模式2：默认关闭当前页（原有逻辑）
            if self._anchor_page is not None:
                try:
                    if not self._anchor_page.is_closed() and old is self._anchor_page:
                        current_page[0] = self._anchor_page
                        await self._settle(current_page[0])
                        return None
                except Exception:
                    pass
            # 硬保护：若当前已是最后一个标签页（通常是列表页/锚点页），
            # 关闭它会导致浏览器窗口整个关掉，必须保留，直接忽略这次关闭。
            try:
                if len(old.context.pages) <= 1:
                    await self._settle(old)
                    return None
            except Exception:
                pass
            try:
                await old.close()
            except Exception:
                pass
            pages = [p for p in old.context.pages if p is not old]
            await self._close_pick_current(pages, None, current_page)
            await self._settle(current_page[0])
            return None
        if atype == "foreach":
            # 通用遍历列表：把当前页匹配 selector 的链接逐个打开→查看→关闭→
            # 全部完成后点 next_selector 翻到下一页继续，直到没有下一页为止。
            # 用于「把公告标题下所有链接都点一遍」这类未知数量、带分页的遍历场景。
            return await self._run_foreach(page, action, current_page, ctx, context_vars)

        if atype == "foreach_if":
            # 逐条检查：遍历列表项，命中 match_text 的才处理（点 click_selector 或
            # 打开该项链接），未命中的自动跳过；全部检查完后再翻页继续。
            # 用于「从表单里逐条检查，命中'确认'、其他跳过」这类按条件筛选的处理场景。
            return await self._run_foreach_if(page, action, current_page, ctx, context_vars)

        if atype == "set_var":
            var = action.get("var", "")
            op = action.get("op", "set")  # set / inc / dec
            raw = action.get("value", 0)
            try:
                num = float(raw)
            except (TypeError, ValueError):
                num = 0
            if float(num).is_integer():
                num = int(num)
            cur = context_vars.get(var, 0)
            if op == "inc":
                base = cur if isinstance(cur, (int, float)) else 0
                val = base + num
                context_vars[var] = int(val) if float(val).is_integer() else val
            elif op == "dec":
                base = cur if isinstance(cur, (int, float)) else 0
                val = base - num
                context_vars[var] = int(val) if float(val).is_integer() else val
            else:  # set
                context_vars[var] = num if isinstance(num, (int, float)) else raw
            return context_vars[var]
        raise ValueError(f"未知动作类型: {atype}")

    # ==== OCR / 结构化抽取 / 报表导出 ====
    async def _action_ocr(self, page: Page, action: Dict[str, Any]) -> str:
        """OCR 读图：整页或元素截图 → 多模态模型识别文字。"""
        sel = (action.get("selector") or "").strip()
        ocr_source = (action.get("ocr_source") or "element").strip()
        try:
            if sel and ocr_source == "element":
                loc = page.locator(sel).first
                if await loc.count() > 0:
                    shot = await loc.screenshot(type="png")
                else:
                    shot = await page.screenshot(type="png")
            else:
                shot = await page.screenshot(type="png")
        except Exception:
            shot = await page.screenshot(type="png")
        return await LLMClient.ocr(shot)

    async def _action_llm_extract(
        self, page: Page, action: Dict[str, Any], context_vars: Dict[str, Any],
        ctx: HealingContext,
    ) -> Dict[str, Any]:
        """结构化 NLP 抽取：从页面/变量/直接文字里按 fields 抽 JSON。"""
        fields = [f.strip() for f in (action.get("fields") or "").replace("，", ",").split(",") if f.strip()]
        # 输入来源：selector（页面） / var（变量） / text（直接文字）
        sel = (action.get("selector") or "").strip()
        var = (action.get("var") or "").strip()
        direct = (action.get("text") or "").strip()
        source_text = ""
        if var:
            source_text = str(context_vars.get(var) or "")
        elif direct:
            source_text = direct
        elif sel:
            try:
                source_text = (await self.operator.extract(
                    page, sel, "text", "", f"提取文字交给 AI 抽取", ctx
                )) or ""
            except Exception:
                source_text = ""
        if not source_text.strip():
            return {}
        # PII 脱敏网关：进 AI 前掩码，返回后还原（外部永远存明文）
        if PIIGateway.is_enabled():
            masked, mapping = PIIGateway.mask(source_text)
        else:
            masked, mapping = source_text, {}
        result = await LLMClient.extract_fields(masked, fields)
        if mapping:
            result = PIIGateway.unmask_dict(result, mapping)
        return result

    def _action_export(
        self, action: Dict[str, Any], context_vars: Dict[str, Any], task_name: str = "",
    ) -> str:
        """把收集到的报表数据导出为 CSV/JSON/Excel/Word/PDF 文件，返回文件路径。

        xlsx/docx 支持引用模板中心模板（template_file）做占位符填充。
        """
        table = context_vars.get("__table") or []
        fmt = (action.get("export_format") or "csv").lower()
        filename = (action.get("export_filename") or task_name or "报表")[:40] or "报表"
        ensure_dirs()
        stamp = time.strftime("%Y%m%d_%H%M%S")
        if not isinstance(table, list) or not table:
            # 空报表也导出带表头的空文件，方便用户确认
            cols = []
            rows = []
        else:
            cols = list({k for r in table for k in r.keys()})
            rows = table
        try:
            if fmt == "json":
                path = EXPORT_DIR / f"{filename}_{stamp}.json"
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(rows, f, ensure_ascii=False, indent=2)
            elif fmt in ("xlsx", "docx", "pdf"):
                template_ref = (action.get("template_file") or "").strip()
                path = EXPORT_DIR / f"{filename}_{stamp}.{fmt}"
                template_path = TemplateEngine.find_template(template_ref) if template_ref else None
                # 标量占位符用首行数据填充（如 {{街道}}←首行的「街道」列），{{__rows__}} 展开整表
                scalar_data = rows[0] if rows else {}
                if fmt == "xlsx":
                    if template_path is not None:
                        TemplateEngine.fill_excel(template_path, scalar_data, rows, path)
                    else:
                        TemplateEngine.to_xlsx(rows, path)
                elif fmt == "docx":
                    if template_path is not None:
                        TemplateEngine.fill_word(template_path, scalar_data, rows, path)
                    else:
                        TemplateEngine.to_docx(rows, path)
                else:
                    TemplateEngine.to_pdf(rows, path)
            else:  # csv
                path = EXPORT_DIR / f"{filename}_{stamp}.csv"
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=cols)
                    writer.writeheader()
                    for r in rows:
                        writer.writerow({k: r.get(k, "") for k in cols})
        except Exception as e:
            raise RuntimeError(f"导出失败: {e}")
        return str(path)

    # ==== 报表数据阶段动作：文件解析 / 清洗 / AI 总结 ====
    def _resolve_file_path(self, raw: str) -> Path:
        """解析本地文件路径：绝对路径直接读，相对路径限定数据目录内。"""
        raw = (raw or "").strip()
        if not raw:
            raise ValueError("缺少文件路径（file_path）")
        p = Path(raw)
        if p.is_absolute():
            if not p.exists():
                raise FileNotFoundError(f"文件不存在: {p}")
            return p
        from config import DATA_DIR

        p = DATA_DIR / raw
        if not p.exists():
            raise FileNotFoundError(f"文件不存在（相对路径按数据目录查找）: {p}")
        return p

    def _action_read_excel(self, action: Dict[str, Any]) -> List[Dict[str, Any]]:
        """读取本地 Excel 文件为表格数据（[{列: 值}, ...]）。"""
        from openpyxl import load_workbook

        path = self._resolve_file_path(action.get("file_path") or "")
        sheet = (action.get("sheet_name") or "").strip() or None
        has_header = bool(action.get("has_header", True))
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            rows_data.append(list(row))
        if not rows_data:
            return []
        if has_header:
            header = [str(h).strip() if h is not None else "" for h in rows_data[0]]
            out = []
            for r in rows_data[1:]:
                out.append({h: (r[i] if i < len(r) else "") for i, h in enumerate(header)})
            return out
        out = []
        for r in rows_data:
            out.append({chr(65 + i): v for i, v in enumerate(r)})
        return out

    def _action_read_csv(self, action: Dict[str, Any]) -> List[Dict[str, Any]]:
        """读取本地 CSV 文件为表格数据。"""
        path = self._resolve_file_path(action.get("file_path") or "")
        encoding = (action.get("encoding") or "utf-8-sig").strip()
        delimiter = (action.get("delimiter") or ",").strip()
        has_header = bool(action.get("has_header", True))
        with open(path, "r", encoding=encoding, newline="") as f:
            rows_data = [row for row in csv.reader(f, delimiter=delimiter) if any((c or "").strip() for c in row)]
        if not rows_data:
            return []
        if has_header:
            header = [h.strip() for h in rows_data[0]]
            out = []
            for r in rows_data[1:]:
                out.append({h: (r[i] if i < len(r) else "") for i, h in enumerate(header)})
            return out
        out = []
        for r in rows_data:
            out.append({chr(65 + i): v for i, v in enumerate(r)})
        return out

    async def _action_ocr_to_json(
        self, page: Page, action: Dict[str, Any], context_vars: Dict[str, Any],
        ctx: HealingContext,
    ) -> List[Dict[str, Any]]:
        """截图 → OCR → 按字段整理成结构化表格（JSON 数组）。"""
        fields = [f.strip() for f in (action.get("fields") or "").replace("，", ",").split(",") if f.strip()]
        text = await self._action_ocr(page, action)
        if not text.strip():
            return []
        if PIIGateway.is_enabled():
            masked, mapping = PIIGateway.mask(text)
        else:
            masked, mapping = text, {}
        rows = await LLMClient.rows_from_text(masked, fields)
        if mapping and rows:
            rows = PIIGateway.unmask_dict(rows, mapping)
        if action.get("append_to_table") and context_vars is not None and rows:
            table = context_vars.get("__table")
            if not isinstance(table, list):
                table = []
            table.extend(rows)
            context_vars["__table"] = table
        return rows

    def _action_data_clean(
        self, action: Dict[str, Any], context_vars: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """数据清洗：去重、空值填充、日期统一、删列。

        flag_missing 命中的关键列缺失 → 标异常并入人工审核队列。
        """
        source = (action.get("source") or "").strip()
        data = context_vars.get(source) if source else context_vars.get("__table")
        if not isinstance(data, list):
            raise ValueError(f"变量 {source or '__table'} 不是表格数据")
        rows = [dict(r) for r in data if isinstance(r, dict)]
        rules = action.get("rules") or {}

        # 去重
        dedup = rules.get("dedup") or []
        if dedup:
            seen = set()
            out = []
            for r in rows:
                key = tuple(str(r.get(k, "")) for k in dedup)
                if key not in seen:
                    seen.add(key)
                    out.append(r)
            rows = out
        # 删列
        for k in rules.get("drop_columns") or []:
            for r in rows:
                r.pop(k, None)
        # 去空白
        if rules.get("trim", True):
            for r in rows:
                for k, v in list(r.items()):
                    if isinstance(v, str):
                        r[k] = v.strip()
        # 空值填充
        for k, v in (rules.get("fill_empty") or {}).items():
            for r in rows:
                if _is_empty(r.get(k)):
                    r[k] = v
        # 日期统一 YYYY-MM-DD
        for k in rules.get("date_format") or []:
            for r in rows:
                if isinstance(r.get(k), str):
                    nd = _normalize_date(r[k])
                    if nd:
                        r[k] = nd
        # 关键列缺失 → 异常入审核队列
        for k in rules.get("flag_missing") or []:
            for r in rows:
                if _is_empty(r.get(k)):
                    r.setdefault("compliance_issues", []).append(f"关键字段缺失：{k}")
                    r["review_status"] = "pending"
        flagged = [r for r in rows if r.get("review_status") == "pending"]
        for r in flagged:
            add_review(
                source="data_clean",
                raw_data=r,
                ai_result={"compliance_issues": r.get("compliance_issues", [])},
                compliance_issues=r.get("compliance_issues", []),
                task_id=context_vars.get("__task_id", ""),
                execution_id=context_vars.get("__exec_id", ""),
            )
        if not source and context_vars is not None:
            context_vars["__table"] = rows
        return rows

    async def _action_llm_summarize(
        self, action: Dict[str, Any], context_vars: Dict[str, Any], ctx: HealingContext,
    ) -> Dict[str, Any]:
        """AI 语义总结 + 异常预警。

        数据进 AI 前脱敏、返回后还原；低置信度 / 命中合规问题 / 异常 → 入人工审核队列。
        """
        from database import get_setting

        source = (action.get("source") or "").strip()
        data = context_vars.get(source) if source else context_vars.get("__table")
        if not isinstance(data, list) or not data:
            return {"overall_summary": "", "rows": [], "flagged_count": 0, "issues": []}
        rows = [dict(r) for r in data if isinstance(r, dict)]
        if not rows:
            return {"overall_summary": "", "rows": [], "flagged_count": 0, "issues": []}
        batch_size = max(1, int(action.get("batch_size") or 10))
        threshold = action.get("threshold")
        if threshold is None:
            try:
                threshold = float(get_setting("review_threshold") or 0.75)
            except (TypeError, ValueError):
                threshold = 0.75
        # 常见姓名列：整值做姓名掩码
        name_fields = ["姓名", "负责人", "经办人", "申请人", "审核人", "户主", "联系人"]
        summary_rows: List[Dict[str, Any]] = []
        overall = ""
        issues: List[str] = []
        for start in range(0, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            # 确定性兜底：数值离群检测（AI 漏检时也能抓住明显异常，强制进人工审核）
            outlier_flags = _detect_numeric_outliers(batch, name_fields)
            if PIIGateway.is_enabled():
                masked_json, mapping = PIIGateway.mask_json_rows(batch, name_fields)
            else:
                masked_json, mapping = json.dumps(batch, ensure_ascii=False), {}
            res = await LLMClient.summarize_rows(masked_json)
            if not overall and res.get("overall_summary"):
                overall = str(res["overall_summary"])
                if mapping:
                    overall, _m = PIIGateway.unmask(overall, mapping)
            covered: set = set()
            for item in res.get("rows") or []:
                if not isinstance(item, dict):
                    continue
                idx = item.get("index")
                try:
                    ridx = int(idx) if idx is not None else None
                except (TypeError, ValueError):
                    ridx = None
                row = rows[ridx] if ridx is not None and 0 <= ridx < len(rows) else None
                if row is None:
                    continue
                covered.add(ridx)
                summary = str(item.get("summary") or "")
                try:
                    conf = float(item["confidence"]) if item.get("confidence") is not None else None
                except (TypeError, ValueError):
                    conf = None
                c_issues = [str(x) for x in (item.get("compliance_issues") or [])]
                anomaly = bool(item.get("anomaly"))
                if mapping:
                    summary, _m = PIIGateway.unmask(summary, mapping)
                    c_issues = [PIIGateway.unmask(x, mapping)[0] for x in c_issues]
                if ridx in outlier_flags:  # 确定性离群 → 无论 AI 是否发现都标异常
                    c_issues = c_issues + outlier_flags[ridx]
                    anomaly = True
                low_conf = conf is not None and conf < threshold
                need_review = low_conf or bool(c_issues) or anomaly
                if row is not None:
                    row["ai_summary"] = summary
                    row["confidence"] = conf
                    row["compliance_issues"] = c_issues
                    row["anomaly"] = anomaly
                    row["review_status"] = "pending" if need_review else "approved"
                    if need_review:
                        issues.extend(c_issues)
                        add_review(
                            source="llm_summarize",
                            raw_data=row,
                            ai_result={
                                "summary": summary, "confidence": conf,
                                "compliance_issues": c_issues, "anomaly": anomaly,
                            },
                            confidence=conf,
                            compliance_issues=c_issues,
                            task_id=context_vars.get("__task_id", ""),
                            execution_id=context_vars.get("__exec_id", ""),
                        )
                summary_rows.append({
                    "index": ridx, "summary": summary, "confidence": conf,
                    "compliance_issues": c_issues, "anomaly": anomaly,
                })
            # 回填：模型漏掉的输入行，标记为「未发现异常」避免报表缺列
            for ridx in range(len(batch)):
                if ridx in covered:
                    continue
                row = rows[ridx]
                row.setdefault("ai_summary", "（未发现异常）")
                row.setdefault("confidence", None)
                row.setdefault("compliance_issues", [])
                row.setdefault("anomaly", False)
                row.setdefault("review_status", "approved")
                if ridx in outlier_flags:  # 离群行即使被模型漏掉也标异常
                    row["compliance_issues"] = row["compliance_issues"] + outlier_flags[ridx]
                    row["anomaly"] = True
                    row["review_status"] = "pending"
                    issues.extend(outlier_flags[ridx])
                    add_review(
                        source="llm_summarize",
                        raw_data=row,
                        ai_result={"summary": "（未发现异常，但数值离群）",
                                   "confidence": None, "compliance_issues": outlier_flags[ridx],
                                   "anomaly": True},
                        compliance_issues=outlier_flags[ridx],
                        task_id=context_vars.get("__task_id", ""),
                        execution_id=context_vars.get("__exec_id", ""),
                    )
                summary_rows.append({
                    "index": ridx, "summary": row["ai_summary"], "confidence": row["confidence"],
                    "compliance_issues": row["compliance_issues"], "anomaly": row["anomaly"],
                })
        if not source and context_vars is not None:
            context_vars["__table"] = rows
        return {
            "overall_summary": overall,
            "rows": summary_rows,
            "flagged_count": len([r for r in summary_rows if r["anomaly"]]),
            "issues": sorted(set(issues)),
        }

    async def _collect_from_page(
        self, detail_page: Page, steps: List[Dict[str, Any]], ctx: HealingContext,
        context_vars: Dict[str, Any],
    ) -> Optional[Dict[str, Any]]:
        """在详情页依次执行提取步骤，返回一行 {save_as: 值}。"""
        row: Dict[str, Any] = {}
        for st in steps:
            if not isinstance(st, dict):
                continue
            stype = st.get("type", "extract")
            save_as = st.get("save_as") or stype
            try:
                if stype == "ocr":
                    val = await self._action_ocr(detail_page, st)
                elif stype == "llm_extract":
                    val = await self._action_llm_extract(detail_page, st, context_vars, ctx)
                    if isinstance(val, dict):
                        for k, v in val.items():
                            row[k] = v if v not in (None, "") else ""
                        continue  # 多字段直接并入行，不再按 save_as 单列
                else:  # extract
                    val = await self.operator.extract(
                        detail_page, st.get("selector", ""),
                        st.get("extract_type", "text"),
                        st.get("attribute", ""), f"提取数据", ctx,
                    )
                row[save_as] = val if val not in (None, "") else ""
            except Exception:
                row[save_as] = ""
        return row

    async def _run_foreach(
        self, page: Page, action: Dict[str, Any], current_page: List[Page], ctx: HealingContext,
        context_vars: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        """通用遍历列表（foreach）。

        在当前（列表）页找到所有匹配 selector 的链接，逐个：
          新标签页打开该链接 → 稍等查看 → 关闭该标签页回到列表页。
        当前页所有链接处理完后，点击 next_selector（"下一页"）刷新列表继续，
        直到没有下一页为止。返回统计信息（打开数量/翻页数）。

        列表页（如巨潮资讯）的数据由 AJAX 异步加载，首次可能为空，
        因此收集链接前会先等待列表出现；翻页后也会等新列表加载完成。
        """
        sel = action.get("selector", "")
        next_sel = action.get("next_selector", "")
        opened = 0
        pages = 0
        max_pages = 50
        try:
            while pages < max_pages:
                pages += 1
                # 收集当前列表页上所有链接的绝对 href（去重、按文档顺序，等待 AJAX 加载）
                hrefs = await self._collect_foreach_hrefs(page, sel)
                if not hrefs:
                    # 首轮就一个都没找到 → 选择器可能错了/页面根本不是列表页，
                    # 直接静默当成功会让用户误以为打开了文件（实际什么都没做），故报错。
                    if pages == 1:
                        raise ElementNotFoundError(self._foreach_sel_err(sel))
                    break
                count = 0
                # 本页是否有「每个详情页提取内容汇总成报表」的需求
                collect_steps = action.get("extract_steps") or []
                record = bool(collect_steps)
                for href in hrefs:
                    new_page = None
                    row = None
                    try:
                        new_page = await page.context.new_page()
                        # 打开详情页（网络偶发失败自动重试一次）
                        ok = await self._safe_goto(new_page, href, attempts=2)
                        if ok:
                            await self._settle(new_page, 0.8)
                            count += 1
                        if ok and record:
                            # 在详情页依次执行提取步骤，汇总成一行的各字段
                            row = await self._collect_from_page(new_page, collect_steps, ctx, context_vars)
                            if context_vars is not None:
                                # 兼容 foreach 直接传 extraction（每个字段各自 save_as 到报表行）
                                pass
                    except Exception:
                        pass
                    finally:
                        if new_page is not None:
                            try:
                                await new_page.close()
                            except Exception:
                                pass
                    if row is not None and context_vars is not None:
                        row["$_href"] = href
                        figure = context_vars.get("__table")
                        if not isinstance(figure, list):
                            figure = []
                        figure.append(row)
                        context_vars["__table"] = figure
                    # 恢复正常浏览：确保回到列表页再开下一个
                    if current_page is not None:
                        try:
                            if current_page[0].is_closed():
                                current_page[0] = page
                        except Exception:
                            current_page[0] = page
                opened += count
                # 全部看完，翻到下一页
                if not next_sel:
                    break
                clicked = False
                try:
                    nxt = page.locator(next_sel).first
                    if await nxt.count() > 0:
                        await nxt.click(timeout=8000)
                        clicked = True
                except Exception:
                    clicked = False
                if not clicked:
                    break
                await self._settle(page)
        except ElementNotFoundError:
            raise  # 首轮无列表链接 → 明确报错，不静默成功
        except Exception:
            pass
        return {"foreach_opened": opened, "foreach_pages": pages}

    async def _collect_foreach_hrefs(self, page: Page, sel: str, max_wait: float = 15.0) -> List[str]:
        """等待列表页 AJAX 加载完成后，收集匹配 selector 的链接 href。

        跳过 about:blank / javascript: 等无效链接。最多等待 max_wait 秒。
        """
        deadline = time.time() + max_wait
        while time.time() < deadline:
            try:
                hrefs = await page.evaluate(
                    """(sel) => {
                        const els = document.querySelectorAll(sel);
                        const out = [];
                        const seen = new Set();
                        for (const el of els) {
                            const a = el.closest('a[href]') || (el.tagName === 'A' ? el : null);
                            if (!a) continue;
                            const h = (a.href || '').trim();
                            if (!h || h === 'about:blank' || h.startsWith('javascript:')) continue;
                            if (!seen.has(h)) { seen.add(h); out.push(h); }
                        }
                        return out;
                    }""",
                    sel,
                )
            except Exception:
                hrefs = []
            if hrefs:
                return hrefs
            await asyncio.sleep(2)
        return []

    async def _run_foreach_if(
        self, page: Page, action: Dict[str, Any], current_page: List[Page], ctx: HealingContext,
        context_vars: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        """逐条检查（foreach_if）。

        在当前（列表）页逐条扫描匹配 selector 的列表项：
          - 命中 match_text（该项文字包含该关键词）的项才处理：
              * 有 click_selector → 点该项内的该按钮/链接（如「确认」）
              * 无 click_selector → 新标签页打开该项链接，看完后关闭
          - 未命中的项自动跳过，不做任何操作。
        当前页全部检查完后，点击 next_selector（"下一页"）继续，直到没有下一页。
        返回统计信息（检查数/命中数/翻页数）。
        """
        sel = action.get("selector", "")
        match_text = action.get("match_text", "")
        click_sel = action.get("click_selector", "")
        next_sel = action.get("next_selector", "")
        checked = matched = pages = 0
        max_pages = 50
        try:
            while pages < max_pages:
                pages += 1
                items = await self._collect_foreach_items(page, sel)
                if not items:
                    # 首轮就一个都没有 → 选择器错 / 页面不是列表页 / 关键字不匹配，
                    # 静默当成功会误导用户以为处理过了，故报错。
                    if pages == 1:
                        raise ElementNotFoundError(self._foreach_sel_err(sel))
                    break
                i = 0
                while True:
                    # 每次重新拉取：上一步操作（如点确认）可能移除/改变了行
                    cur = await self._collect_foreach_items(page, sel)
                    if not cur or i >= len(cur):
                        break
                    item = cur[i]
                    checked += 1
                    text = item.get("text") or ""
                    # 未命中：跳过这条
                    if match_text and match_text not in text:
                        i += 1
                        continue
                    matched += 1
                    if click_sel:
                        # 命中 → 点该项内指定按钮/链接
                        # 记录点击前已有标签页，避免误关用户/其它标签页
                        try:
                            before_ids = {id(p) for p in page.context.pages}
                        except Exception:
                            before_ids = set()
                        await self._click_foreach_target(page, sel, i, click_sel)
                        # 点击后可能有弹窗/新标签页/页面变化，等它稳定
                        await self._settle(current_page[0], 0.6)
                        # 只关闭本次新打开的标签页（如详情链接），回到列表页
                        try:
                            for p in page.context.pages:
                                if id(p) in before_ids:
                                    continue
                                try:
                                    if not p.is_closed() and (p.url or "").strip() not in ("", "about:blank"):
                                        await p.close()
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    else:
                        # 命中 → 打开该项链接，看完后关闭，并尝试按 extract_steps 提取内容入报表
                        href = item.get("href") or ""
                        if href:
                            new_page = None
                            collect_steps = action.get("extract_steps") or []
                            try:
                                new_page = await page.context.new_page()
                                ok = await self._safe_goto(new_page, href, attempts=2)
                                if ok:
                                    await self._settle(new_page, 0.8)
                                if ok and collect_steps:
                                    row = await self._collect_from_page(new_page, collect_steps, ctx, context_vars or {})
                                    if row and context_vars is not None:
                                        row["$_href"] = href
                                        figure = context_vars.get("__table")
                                        if not isinstance(figure, list):
                                            figure = []
                                        figure.append(row)
                                        context_vars["__table"] = figure
                            except Exception:
                                pass
                            finally:
                                if new_page is not None:
                                    try:
                                        await new_page.close()
                                    except Exception:
                                        pass
                    # 恢复当前页为列表页，再处理下一条
                    if current_page is not None:
                        try:
                            if current_page[0].is_closed():
                                current_page[0] = page
                        except Exception:
                            current_page[0] = page
                    # 若行被移除（数量减少），索引保持让后面的项前移补齐；否则前进
                    after = await self._collect_foreach_items(page, sel)
                    if not after:
                        break
                    if len(after) < len(cur):
                        pass
                    else:
                        i += 1
                # 本页检查完，翻下一页
                if not next_sel:
                    break
                clicked = False
                try:
                    nxt = page.locator(next_sel).first
                    if await nxt.count() > 0:
                        await nxt.click(timeout=8000)
                        clicked = True
                except Exception:
                    clicked = False
                if not clicked:
                    break
                await self._settle(page)
        except ElementNotFoundError:
            raise  # 首轮无列表项 → 明确报错，不静默成功
        except Exception:
            pass
        return {"checked": checked, "matched": matched, "foreach_pages": pages}

    def _foreach_sel_err(self, sel: str) -> str:
        """生成 foreach 首轮找不到任何链接时的友好报错。若选择器是中文描述
        （模型把「财务报告下的文件链接」这种文字当成了选择器），给出专门提示。"""
        if re.search(r"[\u4e00-\u9fff]", sel):
            return (
                f"无法定位列表中的文件链接：选择器「{sel}」看起来是文字描述而不是网页元素标识。"
                "请在方案里把这一项改成网页上真实的链接标识（例如 li a），或手动编辑这一步骤。"
            )
        return f"在当前页面没有找到列表文件链接（选择器：{sel}），任务未打开任何文件。请检查打开的网址是否正确，或修改该步骤的链接标识。"

    async def _collect_foreach_items(self, page: Page, sel: str, max_wait: float = 15.0) -> List[Dict[str, Any]]:
        """等待列表页 AJAX 加载完成后，收集匹配 selector 的每个列表项的文本与链接。

        返回 [{text, href}, ...]，按文档顺序。最多等待 max_wait 秒。
        """
        deadline = time.time() + max_wait
        while time.time() < deadline:
            try:
                items = await page.evaluate(
                    """(sel) => {
                        const els = document.querySelectorAll(sel);
                        const out = [];
                        for (const el of els) {
                            const a = el.closest('a[href]') || (el.tagName === 'A' ? el : null);
                            let h = '';
                            if (a) { h = (a.href || '').trim(); }
                            if (!h || h === 'about:blank' || h.startsWith('javascript:')) { h = ''; }
                            out.push({
                                text: (el.innerText || el.textContent || '').trim().slice(0, 300),
                                href: h
                            });
                        }
                        return out;
                    }""",
                    sel,
                )
            except Exception:
                items = []
            if items:
                return items
            await asyncio.sleep(2)
        return []

    async def _click_foreach_target(
        self, page: Page, sel: str, idx: int, click_sel: str
    ) -> bool:
        """点击列表第 idx 项内的目标元素（click_sel 指定的按钮/链接）。

        先在页面上重新定位该项（页面可能已变化），在项内找 click_sel，
        找不到则退回点该项自身；用 nth-child 路径定位目标并做
        普通点击→强制点击→JS 点击 三级兜底。
        """
        try:
            path = await page.evaluate(
                """({sel, idx, clickSel}) => {
                    const els = document.querySelectorAll(sel);
                    const it = els[idx];
                    if (!it) return '';
                    let t = it;
                    if (clickSel) {
                        const c = it.querySelector(clickSel);
                        if (c) t = c;
                    }
                    if (!t) return '';
                    try { t.scrollIntoView({block: 'center', inline: 'nearest'}); } catch (e) {}
                    // 生成从 body 到目标的 nth-child 路径，保证定位唯一
                    const parts = [];
                    let node = t;
                    while (node && node.nodeType === 1 && node.tagName !== 'BODY') {
                        const parent = node.parentElement;
                        let part = node.tagName.toLowerCase();
                        if (parent) {
                            const kids = Array.from(parent.children);
                            const k = kids.indexOf(node);
                            if (k >= 0) part += ':nth-child(' + (k + 1) + ')';
                        }
                        parts.unshift(part);
                        node = parent;
                    }
                    return parts.join(' > ');
                }""",
                {"sel": sel, "idx": idx, "clickSel": click_sel},
            )
        except Exception:
            return False
        if not path:
            return False
        try:
            loc = page.locator(path).first
            if await loc.count() > 0:
                try:
                    await loc.click(timeout=8000)
                except Exception:
                    try:
                        await loc.click(force=True, timeout=5000)
                    except Exception:
                        try:
                            await loc.evaluate("el => el.click()")
                        except Exception:
                            return False
                return True
        except Exception:
            pass
        return False

    async def _close_pick_current(
        self, existing_pages: List[Page], closed_page: Optional[Page], current_page: List[Page]
    ) -> None:
        """close_tab 关闭后，决定新的当前操作页：
        优先回退到「打开被关闭标签前的页面」（标签栈），其次任务起始列表页，再其次非空白页。"""
        old_page = current_page[0] if current_page else None
        # 1) 标签栈回退：s3 从热门页打开视频页 → s4 关闭视频页 → 回到热门页，
        #    保证后续步骤仍在录制时的页面上执行，而不是落到首页。
        prev = self.operator._pop_tab_stack(old_page or closed_page)
        if prev is not None:
            try:
                if prev.is_closed() is False and prev in existing_pages:
                    current_page[0] = prev
                    return
            except Exception:
                pass
        # 2) 栈回退不可用时，回任务起始列表页
        anchor = getattr(self, "_anchor_page", None)
        target = None
        if anchor is not None:
            try:
                if anchor.is_closed() is False and anchor in existing_pages:
                    target = anchor
            except Exception:
                target = None
        if target is None:
            candidates = []
            for p in existing_pages:
                if closed_page is not None and p is closed_page:
                    continue
                try:
                    u = p.url
                except Exception:
                    u = ""
                if u not in ("", "about:blank"):
                    candidates.append(p)
            target = candidates[0] if candidates else None
        if target is not None:
            current_page[0] = target
        elif closed_page is not None:
            # 实在没有可用页面才兜底新开
            try:
                new_p = await closed_page.context.new_page()
                current_page[0] = new_p
            except Exception:
                pass

    async def _settle(self, page: Page, seconds: float = 1.2) -> None:
        """等待页面相对稳定。

        打开页面/点击后会触发动态内容加载（如轮播图动画、热榜请求），
        这里等一次加载态结束并静置片刻，减少下一步定位命中抖动中的元素。
        """
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=6000)
        except Exception:
            pass
        await asyncio.sleep(seconds)

    async def _safe_goto(self, page: Page, url: str, attempts: int = 3) -> bool:
        """带重试的页面导航。

        巨潮资讯等网站会偶发 ERR_EMPTY_RESPONSE / 连接被关闭，一次 goto 可能
        失败或卡在空白页。这里重试数次，返回是否最终成功。
        """
        url = (url or "").strip()
        if not url:
            return False
        for i in range(attempts):
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=20000)
                return True
            except Exception:
                if i >= attempts - 1:
                    return False
                await asyncio.sleep(1.5 + i * 1.0)
        return False

    async def _shot(self, page: Page, dirpath: Path, name: str) -> Optional[str]:
        try:
            path = dirpath / f"{name}.jpg"
            await page.screenshot(path=str(path), type="jpeg", quality=70)
            return str(path)
        except Exception:
            return None